import pandas as pd
import os
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule

# Pfade zu den Excel-Dateien
data_dir = 'survey_2024_2023'
file_2023 = os.path.join(data_dir, 'Mitarbeiterbefragung 2023-T1911-20230702-results-total.xlsx')
file_2024 = os.path.join(data_dir, 'Mitarbeiterbefragung 2024-F2505-Ergebnisexport-Gesamtergebnisse-2024-06-23T235959.xlsx')
output_file = 'Fielmann_MAB_Delta_2023_2024_pro_OrgEinheit_Pivot.xlsx'

def analyze_employee_surveys():
    print("Lese Excel-Dateien...")
    try:
        # Excel-Dateien laden
        excel_2023 = pd.ExcelFile(file_2023)
        excel_2024 = pd.ExcelFile(file_2024)
        
        # Verfügbare Blätter anzeigen
        print(f"Verfügbare Blätter in 2023: {excel_2023.sheet_names}")
        print(f"Verfügbare Blätter in 2024: {excel_2024.sheet_names}")
        
        # Für beide Dateien ist das Blatt 'M' zu verwenden
        sheet_2023 = 'M'
        sheet_2024 = 'M'
        
        # Laden der Daten
        df_2023 = pd.read_excel(file_2023, sheet_name=sheet_2023)
        df_2024 = pd.read_excel(file_2024, sheet_name=sheet_2024)
        
        # Identifiziere alle Items in beiden Datensätzen
        # Die Items sind alle Spalten, die weder die initialen Metadaten-Spalten noch die letzte Spalte sind
        meta_columns_2023 = ['Sort', 'ID', 'Name', 'Level 1', 'Level 2', 'Level 3', 'Level 4', 'Level 5', 
                            'Level 6', 'Level 7', 'Level 8', 'Level 9', 'Level 10', 'Level 11', 'Typ', 'n', 'N']
        meta_columns_2024 = ['Sort', 'ID', 'Name', 'Level 1', 'Level 2', 'Level 3', 'Level 4', 'Level 5', 
                            'Level 6', 'Level 7', 'Level 8', 'Level 9', 'Level 10', 'Level 11', 'Level 12', 'Typ', 'n', 'N']
        
        # Extrahiere Items aus beiden Datensätzen
        items_2023 = [col for col in df_2023.columns if col not in meta_columns_2023]
        items_2024 = [col for col in df_2024.columns if col not in meta_columns_2024]
        
        # Entferne die letzte Spalte von 2023, da es die Umfrage-Frage ist
        if "The staff survey (2022) led to positive changes in my working environment." in items_2023:
            items_2023.remove("The staff survey (2022) led to positive changes in my working environment.")
            
        # Entferne spezielle Spalten aus 2024, die keine direkten Items sind
        special_columns_2024 = ["Information", "Decision making", "Learning from mistakes", "Collaboration", "Leadership", 
                                "The employee survey (2023) led to positive changes in my working environment."]
        items_2024 = [item for item in items_2024 if item not in special_columns_2024]
        
        # Finde gemeinsame Items
        common_items = set(items_2023).intersection(set(items_2024))
        print(f"Anzahl gemeinsamer Items: {len(common_items)}")
        
        # Sammle alle eindeutigen Organisationseinheiten, die in beiden Datensätzen vorkommen
        org_units = []
        for i in range(1, len(df_2023)):
            org_unit_2023 = df_2023.iloc[i]['Name']
            matching_rows_2024 = df_2024[df_2024['Name'] == org_unit_2023]
            if not matching_rows_2024.empty:
                org_units.append(org_unit_2023)
        
        # Erstelle einen neuen DataFrame mit einer Zeile pro Organisationseinheit
        # und Spalten für jedes Item (2023, 2024, Delta)
        result_data = {}
        
        # Füge die Organisationseinheit als erste Spalte hinzu
        result_data['Organisationseinheit'] = sorted(org_units)
        
        # Sortiere die Items für konsistente Reihenfolge
        sorted_items = sorted(common_items)
        
        # Für jedes Item füge drei Spalten hinzu: 2023, 2024, Delta
        for item in sorted_items:
            # Initialisiere die Listen für die Werte
            values_2023 = []
            values_2024 = []
            deltas = []
            
            # Für jede Organisationseinheit hole die Werte
            for org_unit in sorted(org_units):
                org_row_2023 = df_2023[df_2023['Name'] == org_unit].iloc[0]
                org_row_2024 = df_2024[df_2024['Name'] == org_unit].iloc[0]
                
                value_2023 = org_row_2023[item]
                value_2024 = org_row_2024[item]
                
                # Überprüfen, ob die Werte numerisch sind
                if pd.notna(value_2023) and pd.notna(value_2024):
                    delta = value_2024 - value_2023
                else:
                    value_2023 = np.nan
                    value_2024 = np.nan
                    delta = np.nan
                
                values_2023.append(value_2023)
                values_2024.append(value_2024)
                deltas.append(delta)
            
            # Füge die Listen zum Ergebnis-Dictionary hinzu
            short_item_name = item[:30] + '...' if len(item) > 30 else item
            result_data[f"{short_item_name} 2023"] = values_2023
            result_data[f"{short_item_name} 2024"] = values_2024
            result_data[f"{short_item_name} Delta"] = deltas
        
        # Erstelle den DataFrame
        result_df = pd.DataFrame(result_data)
        
        # Formatierung für Excel-Export
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        result_df.to_excel(writer, sheet_name='Delta pro OrgEinheit', index=False)
        
        # Zugriff auf das Arbeitsblatt für Formatierung
        workbook = writer.book
        worksheet = writer.sheets['Delta pro OrgEinheit']
        
        # Spaltenbreiten anpassen
        worksheet.column_dimensions['A'].width = 40  # Organisationseinheit-Spalte
        
        # Anzahl der Items
        num_items = len(sorted_items)
        
        # Für jede Item-Gruppe (3 Spalten) formatieren
        for i in range(num_items):
            # Bestimme die Spaltenbuchstaben für dieses Item
            col_2023 = get_column_letter(i * 3 + 2)  # B, E, H, ...
            col_2024 = get_column_letter(i * 3 + 3)  # C, F, I, ...
            col_delta = get_column_letter(i * 3 + 4)  # D, G, J, ...
            
            # Setze Spaltenbreite
            worksheet.column_dimensions[col_2023].width = 15
            worksheet.column_dimensions[col_2024].width = 15
            worksheet.column_dimensions[col_delta].width = 15
            
            # Bedingte Formatierung für Delta-Spalte
            worksheet.conditional_formatting.add(
                f'{col_delta}2:{col_delta}{len(org_units) + 1}',
                ColorScaleRule(
                    start_type='min',
                    start_color='FF9999',  # Rot für negative Werte
                    mid_type='num',
                    mid_value=0,
                    mid_color='FFFFFF',    # Weiß für 0
                    end_type='max',
                    end_color='99FF99'     # Grün für positive Werte
                )
            )
            
            # Formatiere Zahlen
            for row in range(2, len(org_units) + 2):
                # Konvertiere Spaltenbuchstaben in Zahlenwerte
                col_2023_idx = column_index_from_string(col_2023)
                col_2024_idx = column_index_from_string(col_2024)
                col_delta_idx = column_index_from_string(col_delta)
                
                # Formatiere die Zellen
                for col_idx in [col_2023_idx, col_2024_idx, col_delta_idx]:
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = '0.00'
        
        # Überschriften formatieren
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col in range(1, num_items * 3 + 2):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Rahmen für alle Zellen
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, len(org_units) + 2):
            for col in range(1, num_items * 3 + 2):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Speichern der Datei
        writer._save()
        print(f"Analyse abgeschlossen. Ergebnisse wurden in {output_file} gespeichert.")
        
    except Exception as e:
        print(f"Fehler beim Verarbeiten der Dateien: {e}")
        import traceback
        traceback.print_exc()
        
if __name__ == "__main__":
    analyze_employee_surveys() 